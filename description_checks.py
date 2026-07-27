import hashlib
import os
import re
import unicodedata
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, cast
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from product_attributes import description_to_text


MAX_EXPORT_BYTES = 10_000_000
MAX_EXPORT_UNCOMPRESSED_BYTES = 50_000_000
MAX_EXPORT_ROWS = 5_000
MAX_DESCRIPTION_CHARS = 65_535
REQUIRED_HEADERS = {"id", "ean", "sku", "description", "price", "quantity"}


class DescriptionExportError(ValueError):
    pass


class _DescriptionPreviewParser(HTMLParser):
    BLOCK_TAGS = {
        "div", "p", "section", "article", "h1", "h2", "h3", "h4", "h5", "h6",
        "ul", "ol", "tr",
    }
    SKIP_TAGS = {"script", "style", "template"}

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.parts = []
        self.skip_depth = 0

    def handle_starttag(self, tag, attrs):
        tag = tag.casefold()
        if tag in self.SKIP_TAGS:
            self.skip_depth += 1
        elif not self.skip_depth and tag == "br":
            self.parts.append("\n")
        elif not self.skip_depth and tag == "li":
            self.parts.append("\n• ")
        elif not self.skip_depth and tag in self.BLOCK_TAGS:
            self.parts.append("\n")

    def handle_endtag(self, tag):
        tag = tag.casefold()
        if tag in self.SKIP_TAGS and self.skip_depth:
            self.skip_depth -= 1
        elif not self.skip_depth and (tag == "li" or tag in self.BLOCK_TAGS):
            self.parts.append("\n")

    def handle_data(self, data):
        if not self.skip_depth:
            self.parts.append(data)


def description_preview(value):
    parser = _DescriptionPreviewParser()
    try:
        parser.feed(str(value or ""))
        parser.close()
    except Exception:
        text = str(value or "")
    else:
        text = "".join(parser.parts)
    text = unicodedata.normalize("NFKC", text).replace("\xa0", " ")
    lines = []
    previous_blank = False
    for raw_line in text.splitlines():
        line = re.sub(r"[ \t\r\f\v]+", " ", raw_line).strip()
        if line:
            lines.append(line)
            previous_blank = False
        elif lines and not previous_blank:
            lines.append("")
            previous_blank = True
    return "\n".join(lines).strip()


def normalize_description(value):
    text = description_to_text(value)
    text = unicodedata.normalize("NFKC", text).casefold()
    return re.sub(r"[^\w]+", " ", text, flags=re.UNICODE).strip()


def description_digest(value):
    normalized = normalize_description(value)
    if not normalized:
        return ""
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def description_matches(reference_text, marketplace_value):
    reference = normalize_description(reference_text)
    marketplace = normalize_description(marketplace_value)
    return bool(reference and marketplace and reference in marketplace)


def load_apilo_description_export(path):
    source = Path(path)
    if source.suffix.lower() != ".xlsx":
        raise DescriptionExportError("Eksport Apilo musi być plikiem XLSX.")
    try:
        size = source.stat().st_size
    except OSError as exc:
        raise DescriptionExportError("Nie można odczytać eksportu Apilo.") from exc
    if size < 1 or size > MAX_EXPORT_BYTES:
        raise DescriptionExportError("Eksport Apilo ma nieprawidłowy rozmiar.")
    try:
        with ZipFile(source) as archive:
            if sum(item.file_size for item in archive.infolist()) > MAX_EXPORT_UNCOMPRESSED_BYTES:
                raise DescriptionExportError(
                    "Rozpakowany eksport Apilo przekracza limit 50 MB."
                )
    except BadZipFile as exc:
        raise DescriptionExportError("Nieprawidłowy plik XLSX z Apilo.") from exc

    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise DescriptionExportError("Nieprawidłowy plik XLSX z Apilo.") from exc
    try:
        if len(workbook.sheetnames) != 1:
            raise DescriptionExportError("Eksport Apilo musi zawierać dokładnie jeden arkusz.")
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise DescriptionExportError("Eksport Apilo jest pusty.") from exc
        headers = [str(value or "").strip() for value in raw_headers]
        positions = {name: index for index, name in enumerate(headers) if name}
        missing = sorted(REQUIRED_HEADERS - set(positions))
        if missing:
            raise DescriptionExportError(
                "Eksport Apilo nie zawiera kolumn: " + ", ".join(missing) + "."
            )

        records = []
        seen_ids = set()
        for row_number, row in enumerate(rows, start=2):
            if row_number > MAX_EXPORT_ROWS + 1:
                raise DescriptionExportError("Eksport Apilo przekracza limit 5000 produktów.")
            if not any(value not in (None, "") for value in row):
                continue
            try:
                product_id = int(cast(Any, row[positions["id"]]))
            except (TypeError, ValueError) as exc:
                raise DescriptionExportError(
                    f"Nieprawidłowe ID produktu w wierszu {row_number}."
                ) from exc
            if product_id < 1 or product_id in seen_ids:
                raise DescriptionExportError(
                    f"Powielone lub nieprawidłowe ID produktu w wierszu {row_number}."
                )
            seen_ids.add(product_id)
            description_html = str(row[positions["description"]] or "").strip()
            if len(description_html) > MAX_DESCRIPTION_CHARS:
                raise DescriptionExportError(
                    f"Opis produktu przekracza limit 65535 znaków w wierszu {row_number}."
                )
            description_text = normalize_description(description_html)
            if not description_text:
                raise DescriptionExportError(
                    f"Brak opisu referencyjnego w wierszu {row_number}."
                )
            records.append(
                {
                    "apilo_product_id": product_id,
                    "ean": str(row[positions["ean"]] or "").strip(),
                    "sku": str(row[positions["sku"]] or "").strip(),
                    "description_html": description_html,
                    "description_preview": description_preview(description_html),
                    "description_text": description_text,
                    "description_hash": hashlib.sha256(
                        description_text.encode("utf-8")
                    ).hexdigest(),
                    "export_price": row[positions["price"]],
                    "export_quantity": row[positions["quantity"]],
                }
            )
        if not records:
            raise DescriptionExportError("Eksport Apilo nie zawiera produktów.")
        return records
    finally:
        workbook.close()


def safe_source_name(path):
    return os.path.basename(os.fspath(path))[:200]
